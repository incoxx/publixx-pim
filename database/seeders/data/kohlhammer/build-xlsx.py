# -*- coding: utf-8 -*-
"""Erzeugt docs/features/kohlhammer-datenmodell.xlsx aus dem Manifest.

Zweistufig, weil das Manifest PHP ist:

    php -r '$s = require "database/seeders/data/kohlhammer/structure.php";
             $a = require "database/seeders/data/kohlhammer/attributes.php";
             echo json_encode(["structure"=>$s, "attributes"=>$a], JSON_UNESCAPED_UNICODE);' \
        > /tmp/manifest.json
    KOH_TMP=/tmp python3 database/seeders/data/kohlhammer/build-xlsx.py

Die Mappe ist generiert und wird nach Manifest-Aenderungen neu erzeugt, nicht
von Hand nachgepflegt. Benoetigt openpyxl.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = os.environ.get('KOH_TMP', '.')
d = json.load(open(f'{SP}/manifest.json', encoding='utf-8'))
S, A = d['structure'], d['attributes']

FONT = 'Arial'
H1   = Font(name=FONT, size=14, bold=True)
H2   = Font(name=FONT, size=11, bold=True, color='FFFFFF')
BOLD = Font(name=FONT, size=10, bold=True)
BASE = Font(name=FONT, size=10)
MUTE = Font(name=FONT, size=10, color='808080')
NOTE = Font(name=FONT, size=9, italic=True, color='808080')

HEAD_FILL = PatternFill('solid', fgColor='1F3864')
SUB_FILL  = PatternFill('solid', fgColor='D9E2F3')
CORE_FILL = PatternFill('solid', fgColor='FFF2CC')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)

# Attribute nach Gruppe
by_group = {}
for block, rows in A.items():
    for r in rows:
        by_group.setdefault(r['attribute_type'], []).append((block, r))

group_labels = {g['technical_name']: g['name_de'] for g in S['attribute_types']}
group_order  = [g['technical_name'] for g in S['attribute_types']]

# Kernfelder je Produkttyp
CORE_FOR_PRODUCT = ['RECORDIDENTIFIER', 'PRODUCTEAN', 'DISTINCTIVETITLE', 'PRODUKTTYP', 'PUBLISHINGSTATUS']
CORE_FOR_MASTER  = {
    'contributor': [('sku', 'CTR-{ADR_NR}[-{PERSON_NR}]', 'Leitschlüssel mit Präfix CTR-'),
                    ('name', 'KEYNAMES, NAMESBEFOREKEY', 'Anzeigename des Mitwirkenden')],
    'adresse':     [('sku', 'ADR-{ADR_NR}[-{PERSON_NR}]', 'Leitschlüssel mit Präfix ADR-'),
                    ('name', 'NACHNAME, VORNAME / NAME1', 'Anzeigename der Adresse')],
}

def style_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.border = H2, HEAD_FILL, BORDER
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def title_block(ws, title, subtitle):
    ws['A1'] = title
    ws['A1'].font = H1
    ws['A2'] = subtitle
    ws['A2'].font = NOTE
    return 4

def yn(v):
    return 'ja' if v else ''

# ─── Produkttyp-Tabs ────────────────────────────────────────────────────
TAB_NAMES = {
    'titel': 'Titel',
    'zeitschrift-heft': 'Zeitschrift Heft',
    'zeitschrift-online': 'Zeitschrift Online',
    'zeitschrift-ebook': 'Zeitschrift E-Book',
    'contributor': 'Contributor',
    'adresse': 'Adresse',
}
# Zeilenbereiche der Attributtabellen, für die Zählformeln der Übersicht
attr_ranges = {}

HEADERS = ['Attributgruppe', 'Technischer Name', 'Bezeichnung', 'Datentyp',
           'Werteliste', 'Einheit', 'Übersetzbar', 'Vermehrbar', 'COVER-Spalte']
WIDTHS  = [22, 34, 34, 12, 30, 10, 11, 11, 26]

for pt in S['product_types']:
    tn = pt['technical_name']
    ws = wb.create_sheet(TAB_NAMES[tn])
    is_master = tn in ('contributor', 'adresse')

    row = title_block(
        ws,
        f"Produkttyp: {pt['name_de']}",
        f"technical_name: {tn}   ·   "
        + ('Stammdaten (kein verkaufbares Produkt)' if is_master else 'Titel-/Artikelstammsatz')
        + f"   ·   Preise: {yn(pt.get('has_prices')) or 'nein'}"
        + f"   ·   Medien: {yn(pt.get('has_media')) or 'nein'}"
        + f"   ·   EAN: {yn(pt.get('has_ean')) or 'nein'}"
        + f"   ·   Maße: {yn(pt.get('has_physical_dimensions')) or 'nein'}"
        + (f"   ·   COVER PRODUKTTYP: {pt['cover_produkttyp']}" if pt.get('cover_produkttyp') else ''),
    )

    # Kernfelder
    c = ws.cell(row=row, column=1, value='Kernfelder (feste Spalten der Tabelle products)')
    c.font, c.fill = BOLD, SUB_FILL
    for i in range(2, len(HEADERS) + 1):
        ws.cell(row=row, column=i).fill = SUB_FILL
    row += 1
    style_header(ws, row, ['Spalte', 'COVER-Quelle', 'Hinweis'] + [''] * 6, [22, 34, 60] + [0] * 6)
    row += 1

    if is_master:
        core_rows = CORE_FOR_MASTER[tn]
    else:
        core_rows = [(S['core_fields'][k]['column'], k, S['core_fields'][k]['note'] or '')
                     for k in CORE_FOR_PRODUCT]
    for col, src, note in core_rows:
        for i, v in enumerate([col, src, note], start=1):
            cc = ws.cell(row=row, column=i, value=v)
            cc.font, cc.border, cc.fill = BASE, BORDER, CORE_FILL
            cc.alignment = Alignment(vertical='top', wrap_text=(i == 3))
        row += 1

    row += 1

    # Attribute
    c = ws.cell(row=row, column=1, value='Attribute')
    c.font, c.fill = BOLD, SUB_FILL
    for i in range(2, len(HEADERS) + 1):
        ws.cell(row=row, column=i).fill = SUB_FILL
    row += 1
    style_header(ws, row, HEADERS, WIDTHS)
    row += 1
    first_attr_row = row
    real_count = 0
    open_groups = 0

    groups = pt['default_attribute_groups']
    for g in group_order:
        if g not in groups:
            continue
        entries = by_group.get(g, [])
        if not entries:
            cc = ws.cell(row=row, column=1, value=group_labels[g])
            cc.font, cc.border = BASE, BORDER
            cc2 = ws.cell(row=row, column=2,
                          value='— Attribute entstehen erst aus dem Profiling-Lauf (Schritt 2)')
            cc2.font, cc2.border = MUTE, BORDER
            for i in range(3, len(HEADERS) + 1):
                ws.cell(row=row, column=i).border = BORDER
            row += 1
            open_groups += 1
            continue
        for _block, r in entries:
            vals = [
                group_labels[g],
                r['technical_name'],
                r['name_de'],
                r['data_type'],
                r.get('value_list', ''),
                r.get('default_unit', ''),
                yn(r.get('is_translatable')),
                yn(r.get('is_multipliable')),
                r.get('source') or r.get('note', ''),
            ]
            for i, v in enumerate(vals, start=1):
                cc = ws.cell(row=row, column=i, value=v)
                cc.font, cc.border = BASE, BORDER
                cc.alignment = Alignment(vertical='top', wrap_text=(i in (3, 9)))
            row += 1
            real_count += 1

    attr_ranges[tn] = (TAB_NAMES[tn], real_count, open_groups)
    ws.sheet_view.showGridLines = False

# ─── Beziehungen ────────────────────────────────────────────────────────
ws = wb.create_sheet('Beziehungen')
row = title_block(ws, 'Beziehungstypen',
                  'Ersetzen die PIMCORE-Relationsfelder. Kantenfelder (ObjectMetadata) werden zu Beziehungsattributen.')
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'Bidirektional', 'Quell-Produkttypen',
                       'Ziel-Produkttypen', 'Beziehungsattribute', 'COVER-Quelle'],
             [30, 28, 13, 42, 24, 42, 34])
row += 1
rel_attr = {r['technical_name']: r for r in A['beziehung']}
for rt in S['relation_types']:
    labels = []
    for a in rt['attributes']:
        r = rel_attr.get(a)
        labels.append(f"{a} ({r['data_type']})" if r else a)
    vals = [rt['technical_name'], rt['name_de'], yn(rt['is_bidirectional']) or 'nein',
            ', '.join(rt['source_types']), ', '.join(rt['target_types']),
            '\n'.join(labels) if labels else '—', rt['source']]
    for i, v in enumerate(vals, start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = BASE, BORDER
        cc.alignment = Alignment(vertical='top', wrap_text=True)
    row += 1

row += 2
ws.cell(row=row, column=1, value='Beziehungsattribute im Detail').font = BOLD
row += 1
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'Datentyp', 'Werteliste', 'COVER-Spalte'] + [''] * 2,
             [30, 28, 13, 42, 24, 0, 0])
row += 1
for r in A['beziehung']:
    vals = [r['technical_name'], r['name_de'], r['data_type'], r.get('value_list', ''), r['source']]
    for i, v in enumerate(vals, start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = BASE, BORDER
    row += 1
ws.sheet_view.showGridLines = False

# ─── Preise ─────────────────────────────────────────────────────────────
ws = wb.create_sheet('Preise')
row = title_block(ws, 'Preise',
                  'Land und PRICESTATUS gehören in die Preisart — der Unique-Index auf product_prices führt beides nicht.')
ws.cell(row=row, column=1, value='Preisarten (price_types)').font = BOLD
row += 1
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'COVER PRICETYPECODE', 'COVER PRICESTATUS'] + [''] * 2,
             [22, 36, 26, 20, 0, 0])
row += 1
for pt in S['price_types']:
    vals = [pt['technical_name'], pt['name_de'], pt['cover']['PRICETYPECODE'], pt['cover']['PRICESTATUS']]
    for i, v in enumerate(vals, start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = BASE, BORDER
    row += 1

row += 2
ws.cell(row=row, column=1, value='Kernspalten der Preiszeile (product_prices)').font = BOLD
row += 1
style_header(ws, row, ['Spalte', 'COVER-Quelle', 'Hinweis'] + [''] * 3, [22, 36, 60, 0, 0, 0])
row += 1
for col, src, note in [
    ('amount', 'PRICEAMOUNT', ''),
    ('currency', 'CURRENCYCODE', ''),
    ('country / price_region_id', 'COUNTRYCODE', 'zusätzlich in der Preisart kodiert (Unique-Index)'),
    ('valid_from', 'PRICEEFFECTIVEFROM', 'Teil des Unique-Index'),
    ('valid_to', 'PRICEEFFECTIVEUNTIL', ''),
    ('scale_from', 'MINIMUMORDERQUANTITY', 'Teil des Unique-Index'),
]:
    for i, v in enumerate([col, src, note], start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border, cc.fill = BASE, BORDER, CORE_FILL
        cc.alignment = Alignment(vertical='top', wrap_text=(i == 3))
    row += 1

row += 2
ws.cell(row=row, column=1, value='Preis-Metadaten (neues Feature — je Preiszeile)').font = BOLD
row += 1
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'Werttyp', 'COVER-Spalte'] + [''] * 2,
             [26, 30, 14, 26, 0, 0])
row += 1
for m in S['price_metadata']:
    for i, v in enumerate([m['technical_name'], m['name_de'], m['value_type'], m['source']], start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = BASE, BORDER
    row += 1
ws.sheet_view.showGridLines = False

# ─── Hierarchien ────────────────────────────────────────────────────────
ws = wb.create_sheet('Hierarchien')
row = title_block(ws, 'Hierarchien',
                  'Nur die Wurzeln. Die Knoten entstehen beim Import aus den COVER-Daten.')
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'Typ', 'Quelle / Aufbau'] + [''] * 2,
             [26, 38, 12, 70, 0, 0])
row += 1
for h in S['hierarchies']:
    for i, v in enumerate([h['technical_name'], h['name_de'], h['hierarchy_type'], h['source']], start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = BASE, BORDER
        cc.alignment = Alignment(vertical='top', wrap_text=(i == 4))
    row += 1
ws.sheet_view.showGridLines = False

# ─── Wertelisten ────────────────────────────────────────────────────────
ws = wb.create_sheet('Wertelisten')
row = title_block(ws, 'Wertelisten',
                  'Container sind angelegt. Einträge liefert der Profiling-Lauf (SELECT DISTINCT auf der COVER-Spalte).')
style_header(ws, row, ['Technischer Name', 'Bezeichnung', 'Einträge', 'Belegte Einträge'] + [''] * 2,
             [38, 42, 11, 60, 0, 0])
row += 1
for vl in S['value_lists']:
    entries = vl['entries']
    txt = ', '.join(f"{e['technical_name']} = {e['display_value_de']}" for e in entries) if entries \
        else 'aus Profiling-Lauf'
    for i, v in enumerate([vl['technical_name'], vl['name_de'], len(entries), txt], start=1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.font, cc.border = (BASE if entries else MUTE), BORDER
        cc.alignment = Alignment(vertical='top', wrap_text=(i == 4))
    row += 1
ws.sheet_view.showGridLines = False

# ─── Nicht übernommen ───────────────────────────────────────────────────
ws = wb.create_sheet('Nicht übernommen')
row = title_block(ws, 'Bewusst nicht übernommene PIMCORE-Felder',
                  'Denormalisierungen und im PIMCORE-Import auskommentierte Felder.')
style_header(ws, row, ['PIMCORE-Feld', 'Begründung'] + [''] * 4, [28, 90, 0, 0, 0, 0])
row += 1
for k, v in S['ignored'].items():
    for i, val in enumerate([k, v], start=1):
        cc = ws.cell(row=row, column=i, value=val)
        cc.font, cc.border = BASE, BORDER
        cc.alignment = Alignment(vertical='top', wrap_text=(i == 2))
    row += 1
ws.sheet_view.showGridLines = False

# ─── Übersicht (als erstes Blatt) ───────────────────────────────────────
ws = wb.create_sheet('Übersicht', 0)
row = title_block(
    ws, 'Datenmodell Kohlhammer / COVER in anyPIM',
    'Quelle: database/seeders/data/kohlhammer/ (attributes.php, structure.php) im Repo incoxx/publixx-pim, '
    'Branch claude/kohlhammer-import-analysis-qp8hxd. Abgeleitet aus den PIMCORE-Importskripten '
    'in incoxx/kohlhammer-pimcore:src/AppBundle/Backend/Import/.')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 22

ws.cell(row=row, column=1, value='Produkttypen').font = BOLD
row += 1
style_header(ws, row, ['Produkttyp', 'Bezeichnung', 'Attribute', 'Offene Gruppen', 'Preise', 'Medien', 'EAN', 'Maße', 'COVER PRODUKTTYP'],
             [26, 34, 11, 15, 9, 9, 9, 9, 22])
row += 1
for pt in S['product_types']:
    tn = pt['technical_name']
    _sheet, real_count, open_groups = attr_ranges[tn]
    ws.cell(row=row, column=1, value=tn).font = BASE
    ws.cell(row=row, column=2, value=pt['name_de']).font = BASE
    ws.cell(row=row, column=3, value=real_count).font = BASE
    ws.cell(row=row, column=4, value=open_groups or '—').font = BASE
    ws.cell(row=row, column=5, value=yn(pt.get('has_prices')) or 'nein').font = BASE
    ws.cell(row=row, column=6, value=yn(pt.get('has_media')) or 'nein').font = BASE
    ws.cell(row=row, column=7, value=yn(pt.get('has_ean')) or 'nein').font = BASE
    ws.cell(row=row, column=8, value=yn(pt.get('has_physical_dimensions')) or 'nein').font = BASE
    ws.cell(row=row, column=9, value=pt.get('cover_produkttyp') or '—').font = BASE
    for i in range(1, 10):
        ws.cell(row=row, column=i).border = BORDER
    row += 1

row += 2
ws.cell(row=row, column=1, value='Bestand insgesamt').font = BOLD
row += 1
for label, value in [
    ('Attributgruppen', len(S['attribute_types'])),
    ('Attribute gesamt', sum(len(v) for v in A.values())),
    ('davon Produktattribute', len(A['produkt']) + len(A['produkt_abgeleitet'])),
    ('davon Adresse / E-Mail', len(A['adresse'])),
    ('davon Contributor', len(A['contributor'])),
    ('davon Bearbeiter', len(A['bearbeiter'])),
    ('davon Beziehungsattribute', len(A['beziehung'])),
    ('Produkttypen', len(S['product_types'])),
    ('Beziehungstypen', len(S['relation_types'])),
    ('Preisarten', len(S['price_types'])),
    ('Preis-Metadaten', len(S['price_metadata'])),
    ('Wertelisten', len(S['value_lists'])),
    ('Hierarchien', len(S['hierarchies'])),
]:
    ws.cell(row=row, column=1, value=label).font = BASE
    ws.cell(row=row, column=2, value=value).font = BASE
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2).border = BORDER
    row += 1

row += 2
ws.cell(row=row, column=1, value='Lesehinweise').font = BOLD
row += 1
for line in [
    'Ein Tab je Produkttyp. Jeder Tab zeigt oben die Kernfelder (gelb — feste Spalten der Tabelle products), '
    'darunter die Attribute, gruppiert nach Attributgruppe.',
    'Die Mappe ist vollständig aus dem Manifest generiert (Skript siehe Doku Abschnitt 10) — sie enthält keine Formeln. '
    'Nach einer Manifest-Änderung wird sie neu erzeugt, nicht von Hand nachgepflegt.',
    'Spalte "Offene Gruppen": Attributgruppen, die im Tab als Platzhalterzeile stehen und noch keine Attribute haben.',
    'Die Attributgruppen "Texte" und "Sachgruppen" sind noch leer: ihre Attribute entstehen erst aus dem '
    'Profiling-Lauf gegen die COVER-Datenbank (TEXTTYPECODE, SUBJECTSCHEMENAME).',
    'Datentypen sind kuratiert, nicht aus PIMCORE abgeleitet — dort ist praktisch alles ein Textfeld. '
    'Sie sind am Datenbestand zu verifizieren.',
    'Die Maßeinheiten mm und g sind als ONIX-übliche Annahme gesetzt, ebenfalls zu verifizieren.',
    'Wertelisten ohne belegte Einträge werden aus dem Bestand gefüllt (SELECT DISTINCT auf der COVER-Spalte).',
]:
    c = ws.cell(row=row, column=1, value='· ' + line)
    c.font = BASE
    c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.row_dimensions[row].height = 28
    row += 1

ws.sheet_view.showGridLines = False
ws.freeze_panes = None

out = os.environ.get('KOH_OUT', 'docs/features/kohlhammer-datenmodell.xlsx')
wb.save(out)
print('geschrieben:', out)
print('Tabs:', wb.sheetnames)
